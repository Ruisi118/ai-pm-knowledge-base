#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""幂等回填模板 —— 把 judge 结果写进交付用的标注表。

用法:
    python3 backfill_xlsx.py            # dry-run，只报告不写盘（默认）
    python3 backfill_xlsx.py --write    # 备份后真正写入，写完自动回读比对
    python3 backfill_xlsx.py --verify   # 只回读比对，不写

设计要点（每条都是踩出来的，改脚本时别丢）:
  1. **幂等**：判定会反复改，重跑必须得到同样的表，不能追加、不能错位。
  2. **dry-run 默认**：--write 才落盘，落盘前自动带时间戳备份。
  3. **写前核对表头**：COLS 里的列号与实际表头逐一比对，对不上直接退出。
     表头改过一次就会全表错位，且错位后的数据看起来完全正常。
  4. **写完回读比对**：逐字段读回来和源数据比，ERRORS 必须为 0。
  5. **交付层口径与审计轨迹分离**：因外部原因排除的题（如重复题），
     results/*.json 里保持真实判定不动，只在交付表写排除结论 + review_note 注明。
  6. **非判定行补齐但不伪造**：没走过判定流程的行（第一轮剔除等）补结论类字段，
     n_wrong / spread 等量化判定字段**一律留空**。
"""
import argparse
import datetime
import json
import os
import re
import shutil
import sys

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
RESULTS = os.path.join(BASE, "results")

# ============================ CONFIG（按新任务调整） ============================

XLSX = os.path.join(BASE, "annotation_sheet.xlsx")
SHEET = "第二轮_模型区分性判断"
ID_COL = 1                      # uniq_query_id 所在列（1-based）
HEADER_ROW = 1
FIRST_DATA_ROW = 2

# 表头名 -> 列号（1-based）。写盘前会与实际表头逐一核对。
COLS = {
    "uniq_query_id": 1,
    "n_models_scored": 12,
    "是否入选": 13,
    "选择理由": 14,
    "n_wrong": 15,
    "wrong_models": 16,
    "selection_type": 17,
    "rejection_type": 18,
    "verification_blocked": 19,
    "divergence_source": 20,
    "safety_sensitive": 21,
    "spread": 22,
    "trap_divergence": 23,
    "explicit_check_count": 24,
    "review_note": 25,
    "stage_status": 26,
}

# 交付层排除：判定为入选，但因外部原因不交付（如该陷阱先前数据已收录）
EXCLUDE_FLAG = "duplicate_of_prior_data"
EXCLUDE_REJECTION = "重复query"
EXCLUDE_NOTE = "判定为入选，因先前数据已收录而在交付层排除；判定字段保留真值。"

# ==============================================================================


def norm(v):
    """把 None / 空串归一，便于比对。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return str(v).strip()


def load_results():
    """读全部结果文件 -> {id: row_dict}。"""
    out, seen = {}, {}
    files = sorted(f for f in os.listdir(RESULTS) if re.fullmatch(r"batch_\d+\.json", f))
    for fn in files:
        data = json.load(open(os.path.join(RESULTS, fn), encoding="utf-8"))
        for it in data["items"]:
            qid = it["uniq_query_id"]
            if qid in seen:
                print("  [warn] %s 在 %s 与 %s 重复出现" % (qid, seen[qid], fn))
            seen[qid] = fn
            s1, s3 = it["step1"], it["step3"]
            row = {
                "n_models_scored": s3.get("n_models"),
                "是否入选": s3["selected"],
                "选择理由": s3.get("selection_reason"),
                "n_wrong": s3.get("n_wrong"),
                "wrong_models": ", ".join(s3.get("wrong_models") or []),
                "selection_type": s3.get("selection_type"),
                "rejection_type": s3.get("rejection_type"),
                "verification_blocked": s3.get("verification_blocked"),
                "divergence_source": s3.get("divergence_source"),
                "safety_sensitive": s1.get("safety_sensitive"),
                "spread": s3.get("spread"),
                "trap_divergence": s3.get("trap_divergence"),
                "explicit_check_count": s3.get("explicit_check_count"),
                "review_note": s3.get("review_note"),
                "stage_status": "第二轮-已完成",
            }
            # 交付层排除：判定字段保留真值，只改结论
            if it.get(EXCLUDE_FLAG):
                row["是否入选"] = "否"
                row["selection_type"] = None
                row["divergence_source"] = None
                row["rejection_type"] = EXCLUDE_REJECTION
                row["review_note"] = EXCLUDE_NOTE
            out[qid] = row
    print("读入 %d 个结果文件，%d 题" % (len(files), len(out)))
    return out


def fill_non_adjudicated(qid, ws, r):
    """没走过判定流程的行：补结论类字段，量化判定字段留空。

    按新任务改写。返回 dict 或 None（None = 本行不归本脚本管）。
    典型用法：读第一轮 sheet 的 quality_flag / quality_note，
    拼成 `【淘汰理由】query语义缺陷：` + quality_note 原文（脚本取，不手抄）。
    """
    return None


def verify_headers(ws):
    bad = []
    for name, col in COLS.items():
        actual = norm(ws.cell(HEADER_ROW, col).value)
        if actual != name:
            bad.append("列%d 期望 %r 实际 %r" % (col, name, actual))
    if bad:
        print("表头核对失败，已中止（表头改过会导致全表错位）:")
        for b in bad:
            print("  -", b)
        sys.exit(1)
    print("表头核对通过")


def run(write, verify_only):
    results = load_results()
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    verify_headers(ws)

    planned, missing, errors = {}, [], []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        qid = norm(ws.cell(r, ID_COL).value)
        if not qid:
            continue
        row = results.get(qid) or fill_non_adjudicated(qid, ws, r)
        if row is None:
            missing.append(qid)
            continue
        planned[r] = row

    print("待回填 %d 行；无判定结果 %d 行" % (len(planned), len(missing)))
    if missing:
        print("  前 10 个无结果的 id:", missing[:10])

    if verify_only:
        for r, row in planned.items():
            for name, val in row.items():
                got = norm(ws.cell(r, COLS[name]).value)
                if got != norm(val):
                    errors.append("行%d %s: 表内 %r != 源 %r" % (r, name, got, norm(val)))
        print("VERIFY ERRORS:", len(errors))
        for e in errors[:50]:
            print(" -", e)
        return

    if not write:
        sample = list(planned.items())[:3]
        print("\n[dry-run] 不写盘。样例:")
        for r, row in sample:
            print("  行%d -> %s" % (r, {k: row[k] for k in ("是否入选", "n_wrong", "rejection_type")}))
        print("\n确认无误后加 --write 执行。")
        return

    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = XLSX.replace(".xlsx", "_backup_%s_回填前.xlsx" % stamp)
    shutil.copy2(XLSX, backup)
    print("已备份 ->", os.path.basename(backup))

    for r, row in planned.items():
        for name, val in row.items():
            ws.cell(r, COLS[name]).value = "" if val is None else val
    wb.save(XLSX)
    print("已写入", XLSX)

    # 回读比对
    wb2 = openpyxl.load_workbook(XLSX)
    ws2 = wb2[SHEET]
    for r, row in planned.items():
        for name, val in row.items():
            got = norm(ws2.cell(r, COLS[name]).value)
            if got != norm(val):
                errors.append("行%d %s: 写回 %r != 源 %r" % (r, name, got, norm(val)))
    print("READBACK ERRORS:", len(errors))
    for e in errors[:50]:
        print(" -", e)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true", help="真正写盘（默认 dry-run）")
    ap.add_argument("--verify", action="store_true", help="只回读比对，不写")
    a = ap.parse_args()
    run(a.write, a.verify)
